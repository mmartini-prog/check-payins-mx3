"""
reconciliation.py — Motor de conciliación Banco vs Payins.

Responsabilidades:
  - Agrupar Kyriba y Payins por Processor + Day
  - Calcular diferencias y porcentajes
  - Clasificar estado (OK, alerta, rojo)
  - Generar resúmenes por processor y global
  - Exportar Excel con formato profesional
  - Sin dependencias de Streamlit
"""

from __future__ import annotations

from io import BytesIO
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    Alignment, Font, PatternFill, numbers
)
from openpyxl.utils import get_column_letter


# ── CONSTANTES ─────────────────────────────────────────────────────────────────

STATUS_OK     = "✅ OK"
STATUS_WARN   = "🟡 Banco mayor"
STATUS_RED    = "🔴 Banco menor"
STATUS_NODATA = "⚠️ Sin dato Payins"
STATUS_NOBANK = "⬜ Sin dato Banco"


# ── RECONCILE ─────────────────────────────────────────────────────────────────

def reconcile(
    kyriba_df: pd.DataFrame,
    payins_df: pd.DataFrame,
    tolerance_pct: float = 10.0,
    selected_processors: Optional[list[str]] = None,
) -> dict:
    """
    Ejecuta la conciliación completa.

    Returns:
        dict con claves:
            - detail (DataFrame): fila por Processor + Day
            - summary (DataFrame): fila por Processor
            - kpis (dict): totales globales
            - no_match (DataFrame): filas con estado != OK
    """
    kyr = kyriba_df.copy()
    pay = payins_df.copy()

    if selected_processors is not None:
        kyr = kyr[kyr["Processor"].isin(selected_processors)]
        pay = pay[pay["Processor"].isin(selected_processors)]

    # ── Agrupar ───────────────────────────────────────────────────────────────
    banco_grouped = (
        kyr[kyr["Processor"].notna()]
        .groupby(["Processor", "Day"], as_index=False)["Credit"]
        .sum()
        .rename(columns={"Credit": "Banco"})
    )

    payins_grouped = (
        pay[pay["Processor"].notna()]
        .groupby(["Processor", "Day"], as_index=False)["Amount"]
        .sum()
        .rename(columns={"Amount": "Payins estimados"})
    )

    # ── Merge outer ───────────────────────────────────────────────────────────
    detail = pd.merge(
        banco_grouped,
        payins_grouped,
        on=["Processor", "Day"],
        how="outer",
    ).fillna(0)

    detail["Diferencia"] = detail["Banco"] - detail["Payins estimados"]
    detail["Dif %"] = detail.apply(_pct, axis=1)
    detail["Estado"] = detail.apply(lambda r: _status(r, tolerance_pct), axis=1)
    detail = detail.sort_values(["Processor", "Day"]).reset_index(drop=True)

    # ── Summary por Processor ─────────────────────────────────────────────────
    summary = (
        detail
        .groupby("Processor", as_index=False)
        .agg({"Banco": "sum", "Payins estimados": "sum", "Diferencia": "sum"})
    )
    summary["Dif %"] = summary.apply(_pct, axis=1)
    summary["Estado"] = summary.apply(lambda r: _status(r, tolerance_pct), axis=1)
    summary = summary.sort_values("Processor").reset_index(drop=True)

    # ── KPIs globales ─────────────────────────────────────────────────────────
    total_banco  = summary["Banco"].sum()
    total_payins = summary["Payins estimados"].sum()
    total_diff   = total_banco - total_payins
    total_pct    = (total_diff / total_payins * 100) if total_payins else 0

    kpis = {
        "total_banco":  total_banco,
        "total_payins": total_payins,
        "total_diff":   total_diff,
        "total_pct":    total_pct,
        "n_ok":         (summary["Estado"] == STATUS_OK).sum(),
        "n_warn":       (summary["Estado"] != STATUS_OK).sum(),
    }

    no_match = detail[detail["Estado"] != STATUS_OK].copy()

    return {
        "detail":    detail,
        "summary":   summary,
        "kpis":      kpis,
        "no_match":  no_match,
    }


# ── HELPERS ────────────────────────────────────────────────────────────────────

def _pct(row) -> float:
    if row["Payins estimados"] != 0:
        return row["Diferencia"] / row["Payins estimados"] * 100
    return 0.0


def _status(row, tolerance: float) -> str:
    if row["Banco"] == 0 and row["Payins estimados"] == 0:
        return STATUS_NODATA
    if row["Payins estimados"] == 0:
        return STATUS_NODATA
    if row["Banco"] == 0:
        return STATUS_NOBANK
    if abs(row["Dif %"]) <= tolerance:
        return STATUS_OK
    return STATUS_RED if row["Diferencia"] < 0 else STATUS_WARN


# ── EXPORT EXCEL ───────────────────────────────────────────────────────────────

_HEADER_FILL   = PatternFill("solid", fgColor="0D0D3D")
_OK_FILL       = PatternFill("solid", fgColor="0A3D1A")
_WARN_FILL     = PatternFill("solid", fgColor="3D2D00")
_RED_FILL      = PatternFill("solid", fgColor="3D0A0A")
_NODATA_FILL   = PatternFill("solid", fgColor="1A1A2E")
_HEADER_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
_DATA_FONT     = Font(name="Arial", size=10)
_CENTER        = Alignment(horizontal="center", vertical="center")
_LEFT          = Alignment(horizontal="left", vertical="center")

_STATUS_FILL = {
    STATUS_OK:     _OK_FILL,
    STATUS_WARN:   _WARN_FILL,
    STATUS_RED:    _RED_FILL,
    STATUS_NODATA: _NODATA_FILL,
    STATUS_NOBANK: _NODATA_FILL,
}

_FMT_NUM   = '#,##0'
_FMT_PCT   = '0.0"%"'


def build_excel(
    summary_df:   pd.DataFrame,
    detail_df:    pd.DataFrame,
    kyriba_df:    pd.DataFrame,
    payins_df:    pd.DataFrame,
    unmapped_df:  pd.DataFrame,
    no_match_df:  pd.DataFrame,
    entity:       str,
    date_range:   tuple,
) -> bytes:
    """Construye el Excel de conciliación con formato profesional."""
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Hoja info
        info_data = {
            "Campo": ["Entidad", "Rango fecha", "Generado"],
            "Valor": [entity, f"{date_range[0]} al {date_range[1]}", pd.Timestamp.now().strftime("%Y-%m-%d %H:%M")],
        }
        pd.DataFrame(info_data).to_excel(writer, sheet_name="Info", index=False)

        summary_df.to_excel(writer, sheet_name="Resumen",           index=False)
        detail_df.to_excel(writer,  sheet_name="Detalle",           index=False)
        kyriba_df.to_excel(writer,  sheet_name="Kyriba combinado",  index=False)
        payins_df.to_excel(writer,  sheet_name="Payins combinado",  index=False)
        unmapped_df.to_excel(writer,sheet_name="Kyriba no mapeado", index=False)
        no_match_df.to_excel(writer,sheet_name="No conciliados",    index=False)

    # Post-process con openpyxl para formato
    output.seek(0)
    wb = load_workbook(output)
    _format_sheet(wb["Resumen"],  amount_cols=["Banco","Payins estimados","Diferencia"], pct_cols=["Dif %"])
    _format_sheet(wb["Detalle"],  amount_cols=["Banco","Payins estimados","Diferencia"], pct_cols=["Dif %"])
    _auto_width(wb)

    out2 = BytesIO()
    wb.save(out2)
    out2.seek(0)
    return out2.read()


def _format_sheet(ws, amount_cols: list[str], pct_cols: list[str]):
    """Aplica formato y color a las hojas principales."""
    headers = {cell.value: cell.column for cell in ws[1]}

    # Header row
    for cell in ws[1]:
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = _CENTER

    # Columnas numéricas
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            h = ws.cell(row=1, column=cell.column).value
            if h in amount_cols:
                cell.number_format = _FMT_NUM
            if h in pct_cols:
                cell.number_format = _FMT_PCT
            cell.font      = _DATA_FONT
            cell.alignment = _CENTER

    # Color por estado
    if "Estado" in headers:
        estado_col = headers["Estado"]
        for row in ws.iter_rows(min_row=2):
            status_cell = ws.cell(row=row[0].row, column=estado_col)
            fill = _STATUS_FILL.get(status_cell.value)
            if fill:
                for cell in row:
                    cell.fill = fill


def _auto_width(wb):
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value or "")))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 4, 50)
